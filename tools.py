import datetime
import os
import pathlib
import time
from copy import copy

import xlrd
import xlsxwriter
from openpyxl import load_workbook, Workbook


# 处理路径
def transform_path(path_dir):
    return path_dir.replace('//', '////')


def get_files(path, all_files):
    files = os.listdir(path)
    for file in files:
        cur_path = os.path.join(path, file)
        if os.path.isdir(cur_path):
            get_files(cur_path, all_files)
        else:
            all_files.append(cur_path)


# 获取文件夹下的所有文件
def get_file_list(pathdir):
    if os.path.exists(pathdir):
        file_list = []
        get_files(pathdir, file_list)
        return file_list
    else:
        raise FileNotFoundError


# 根据给定的后缀名数组过滤
def filter_by_condition(arr, external_list):
    filter_list = []
    for item in arr:
        if os.path.splitext(item)[-1] in external_list:
            filter_list.append(item)
    return filter_list


# 格式化日期时间
def format_data(col, book):
    result = col.value
    if col.ctype == 3:
        result = datetime.datetime(*xlrd.xldate_as_tuple(col.value, book.datemode)).strftime('%Y/%m/%d')
    return result


# 读取excel文件
def read_sheet(filepath, exclude_list, content):
    book = xlrd.open_workbook(filepath)
    for sheet in book.sheets():
        if sheet.name not in exclude_list:
            for index, row_data in enumerate(sheet.get_rows()):
                # 去除前三行和最后两行
                if 3 <= index < sheet.nrows - 2:
                    row_content = []
                    for col_data in row_data:
                        row_content.append(format_data(col_data, book))
                    content.append(row_content)
    return content


# 写入sheet
def write_sheet1(content_list):
    excel = xlsxwriter.Workbook('汇总.xlsx')
    book = excel.add_worksheet('汇总数据')
    # 书写title
    title_list = ['目的公司', '目标调入金额', '调出公司', '金额', '中转公司1/目的公司', '金额2', '中转公司2/目的公司', '金额3', '中转公司3/目的公司', '余额更新', '调拨日期', '流程编号', '备注']
    for index, title_data in enumerate(title_list):
        book.write(0, index, title_data)
    for row_index, data in enumerate(content_list):
        for (col_index, row_data) in enumerate(data):
            book.write(row_index+1, col_index, row_data)
    excel.close()


# 写入sheet
def write_sheet(filename, content_list):
    if not pathlib.Path(filename).exists():
        wb = Workbook()
    else:
        wb = load_workbook(filename)
    ws = wb.active
    title_list = ['目的公司', '目标调入金额', '调出公司', '金额', '中转公司1/目的公司', '金额2', '中转公司2/目的公司',
                  '金额3', '中转公司3/目的公司', '余额更新', '调拨日期', '流程编号', '备注']
    for index, title_data in enumerate(title_list):
        ws.cell(1, index+1).value = title_data

    for row_index, row_data in enumerate(content_list):
        ws.row_dimensions[row_index + 1].height = 64.5
        for col_index, col_data in enumerate(row_data):
            ws.cell(row_index+2, col_index + 1).value = col_data
        time.sleep(0.01)
    ws.column_dimensions['A'].width = 31.89 + 0.78
    ws.column_dimensions['B'].width = 16.44 + 0.78
    ws.column_dimensions['C'].width = 31.89 + 0.78
    ws.column_dimensions['D'].width = 16.44 + 0.78
    ws.column_dimensions['E'].width = 30.89 + 0.78
    ws.column_dimensions['F'].width = 14.67 + 0.78
    ws.column_dimensions['G'].width = 29.67 + 0.78
    ws.column_dimensions['H'].width = 16.89 + 0.78
    ws.column_dimensions['I'].width = 29.67 + 0.78
    ws.column_dimensions['J'].width = 8 + 0.78
    ws.column_dimensions['K'].width = 11 + 0.78
    ws.column_dimensions['L'].width = 24.56 + 0.78
    ws.column_dimensions['M'].width = 22.67 + 0.78
    wb.save(filename)


# 复制模板
def copy_template(filename):
    # filename = 'template.xlsx'
    if os.path.exists(filename):
        wb = load_workbook(filename)
        sheet = wb['模板']
        ws = wb.active
        cols = ws.max_column
        rows = ws.max_row
        ws1 = wb.create_sheet('复制的模板')
        for row in range(0, rows):
            for col in range(0, cols):
                ws1.cell(row+1, col+1).value = sheet.cell(row+1, col+1).value
                ws1.cell(row+1, col+1)._style = copy(sheet.cell(row+1, col+1)._style)
                ws1.row_dimensions[row+1].height = 30.0
                ws1.column_dimensions['A'].width = 18.34
                ws1.column_dimensions['B'].width = 30.11
                ws1.column_dimensions['C'].width = 16.45
                ws1.column_dimensions['D'].width = 40.78
        ws1.merge_cells("A1:C1")
        wb.save(filename)
